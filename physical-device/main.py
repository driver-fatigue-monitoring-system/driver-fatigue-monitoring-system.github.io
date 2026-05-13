import os

os.environ.setdefault("PYGAME_HIDE_SUPPORT_PROMPT", "1")

import json
import sys
import time
import threading
from collections import deque
import urllib.error
import urllib.request
import uuid
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path

import cv2

try:
    import pygame
except ImportError:
    pygame = None


ROOT_DIR = Path(__file__).resolve().parent
DEVICE_CONFIG_PATH = Path(os.getenv("DFMS_DEVICE_CONFIG", ROOT_DIR / "device_config.json"))


def load_device_config():
    if not DEVICE_CONFIG_PATH.exists():
        return {}
    try:
        config = json.loads(DEVICE_CONFIG_PATH.read_text(encoding="utf-8"))
        return config if isinstance(config, dict) else {}
    except Exception as error:
        print(f"Could not read device config {DEVICE_CONFIG_PATH}: {error}")
        return {}


DEVICE_CONFIG = load_device_config()
DEFAULT_MODEL_PATH = ROOT_DIR / "fatigues.pt"
MODEL_PATH = Path(
    os.getenv(
        "FATIGUE_MODEL_PATH",
        DEFAULT_MODEL_PATH if DEFAULT_MODEL_PATH.exists() else ROOT_DIR / "models" / "fatigues.pt",
    )
)
SUPABASE_EVENT_URL = (
    DEVICE_CONFIG.get("supabase_event_url")
    or os.getenv("SUPABASE_EVENT_URL")
    or "https://wxxdtuzicvjkjqymjsfj.supabase.co/functions/v1/device-events"
).strip()
DEVICE_SERIAL = (DEVICE_CONFIG.get("device_serial") or os.getenv("DFMS_DEVICE_SERIAL", "")).strip()
DEVICE_TOKEN = (DEVICE_CONFIG.get("device_token") or os.getenv("DFMS_DEVICE_TOKEN", "")).strip()
DEVICE_SESSION_ID = os.getenv("DFMS_SESSION_ID", f"session-{uuid.uuid4()}")
EVENT_SEND_TIMEOUT_SECONDS = float(os.getenv("DFMS_EVENT_SEND_TIMEOUT", "5"))
EVENT_COOLDOWN_SECONDS = float(os.getenv("DFMS_EVENT_COOLDOWN", "5"))
DROWSY_ALARM_PATH = Path(os.getenv("DFMS_DROWSY_ALARM_PATH", ROOT_DIR / "alarm1.mp3"))
YAWN_ALARM_PATH = Path(os.getenv("DFMS_YAWN_ALARM_PATH", ROOT_DIR / "alarm2.mp3"))
DROWSY_ALARM_DELAY_SECONDS = float(os.getenv("DFMS_DROWSY_ALARM_DELAY", "3"))
YAWN_ALARM_COOLDOWN_SECONDS = float(os.getenv("DFMS_YAWN_ALARM_COOLDOWN", "10"))
CAMERA_WIDTH = int(os.getenv("DFMS_CAMERA_WIDTH", "640"))
CAMERA_HEIGHT = int(os.getenv("DFMS_CAMERA_HEIGHT", "480"))
MODEL_IMGSZ = int(os.getenv("DFMS_MODEL_IMGSZ", "0"))
YAWN_MIN_CONFIDENCE = float(os.getenv("DFMS_YAWN_MIN_CONFIDENCE", "0.65"))
DROWSY_RECORDING_DIR = Path(os.getenv("DFMS_DROWSY_RECORDING_DIR", ROOT_DIR / "recordinges_drowsy"))
DROWSY_RECORDING_SECONDS = float(os.getenv("DFMS_DROWSY_RECORDING_SECONDS", "15"))
DROWSY_RECORDING_PRE_SECONDS = float(os.getenv("DFMS_DROWSY_RECORDING_PRE_SECONDS", "3"))
DROWSY_RECORDING_MAX_BYTES = int(float(os.getenv("DFMS_DROWSY_RECORDING_MAX_MB", "1024")) * 1024 * 1024)
DROWSY_RECORDING_CODEC = os.getenv("DFMS_DROWSY_RECORDING_CODEC", "mp4v")
EXCEL_LOG_PATH = Path(os.getenv("DFMS_EXCEL_LOG_PATH", ROOT_DIR / "Excel_driver_logs.xlsx"))
EXCEL_MAX_BYTES = int(float(os.getenv("DFMS_EXCEL_MAX_MB", "100")) * 1024 * 1024)
EXCEL_TRIM_BYTES = int(float(os.getenv("DFMS_EXCEL_TRIM_MB", "30")) * 1024 * 1024)

COLORS = {
    "Awake": (0, 255, 0),
    "Drowsy": (0, 0, 255),
    "Yawn": (0, 165, 255),
    "No Detection": (148, 163, 184),
}

_last_event_sent_at = {}


class AlarmController:
    def __init__(self):
        self.drowsy_started_at = None
        self.drowsy_alarm_active = False
        self.last_yawn_alarm_at = 0.0
        self.audio_ready = False
        self.audio_warning_shown = False
        self.audio_error = ""
        self.current_music = ""
        self.last_status = "Alarm ready"

    def ensure_audio(self):
        if self.audio_ready:
            return True
        if pygame is None:
            self.show_audio_warning("pygame is not installed. Run: pip install pygame")
            return False
        if not DROWSY_ALARM_PATH.exists():
            self.show_audio_warning(f"Missing drowsy alarm file: {DROWSY_ALARM_PATH}")
            return False
        if not YAWN_ALARM_PATH.exists():
            self.show_audio_warning(f"Missing yawn alarm file: {YAWN_ALARM_PATH}")
            return False

        try:
            pygame.mixer.pre_init(frequency=44100, size=-16, channels=2, buffer=512)
            pygame.init()
            if pygame.mixer.get_init() is None:
                pygame.mixer.init(frequency=44100, size=-16, channels=2, buffer=512)
            pygame.mixer.music.set_volume(1.0)
            self.audio_ready = True
            self.audio_error = ""
            self.last_status = "Alarm audio ready"
            return True
        except Exception as error:
            self.show_audio_warning(f"Could not initialize alarm audio: {error}")
            return False

    def show_audio_warning(self, message):
        self.audio_error = message
        self.last_status = message
        print(message)
        self.audio_warning_shown = True

    def update(self, current_status, detections):
        detected_statuses = {detection.get("status") for detection in detections}
        now = time.monotonic()

        if "Drowsy" in detected_statuses or current_status == "Drowsy":
            if self.drowsy_started_at is None:
                self.drowsy_started_at = now
            elif now - self.drowsy_started_at >= DROWSY_ALARM_DELAY_SECONDS:
                self.start_drowsy_alarm()
        else:
            self.drowsy_started_at = None
            self.stop_drowsy_alarm()

        if "Yawn" in detected_statuses or current_status == "Yawn":
            if now - self.last_yawn_alarm_at >= YAWN_ALARM_COOLDOWN_SECONDS:
                self.play_yawn_alarm()
                self.last_yawn_alarm_at = now

        if not self.audio_error:
            self.last_status = self.status_text(now)

    def start_drowsy_alarm(self):
        if self.drowsy_alarm_active:
            return
        if not self.ensure_audio():
            return
        try:
            pygame.mixer.music.load(str(DROWSY_ALARM_PATH))
            pygame.mixer.music.play(loops=-1)
            self.current_music = "Drowsy"
            self.drowsy_alarm_active = True
            self.last_status = "Drowsy alarm ON"
        except Exception as error:
            self.show_audio_warning(f"Could not play drowsy alarm: {error}")

    def stop_drowsy_alarm(self):
        if not self.drowsy_alarm_active:
            return
        pygame.mixer.music.stop()
        self.current_music = ""
        self.drowsy_alarm_active = False

    def play_yawn_alarm(self):
        if not self.ensure_audio():
            return
        if self.drowsy_alarm_active:
            return
        try:
            pygame.mixer.music.load(str(YAWN_ALARM_PATH))
            pygame.mixer.music.play(loops=0)
            self.current_music = "Yawn"
            self.last_status = "Yawn alarm played"
        except Exception as error:
            self.show_audio_warning(f"Could not play yawn alarm: {error}")

    def test_drowsy_alarm(self):
        if not self.ensure_audio():
            return
        try:
            pygame.mixer.music.load(str(DROWSY_ALARM_PATH))
            pygame.mixer.music.play(loops=0)
            self.current_music = "Drowsy test"
            self.last_status = "Testing drowsy alarm"
            print(f"Testing drowsy alarm: {DROWSY_ALARM_PATH}")
        except Exception as error:
            self.show_audio_warning(f"Could not test drowsy alarm: {error}")

    def test_yawn_alarm(self):
        if not self.ensure_audio():
            return
        try:
            pygame.mixer.music.load(str(YAWN_ALARM_PATH))
            pygame.mixer.music.play(loops=0)
            self.current_music = "Yawn test"
            self.last_status = "Testing yawn alarm"
            print(f"Testing yawn alarm: {YAWN_ALARM_PATH}")
        except Exception as error:
            self.show_audio_warning(f"Could not test yawn alarm: {error}")

    def stop_all(self):
        if pygame is not None and pygame.mixer.get_init() is not None:
            pygame.mixer.music.stop()
        self.current_music = ""
        self.drowsy_alarm_active = False

    def status_text(self, now=None):
        now = now or time.monotonic()
        drowsy_seconds = 0.0
        if self.drowsy_started_at is not None:
            drowsy_seconds = max(0.0, now - self.drowsy_started_at)
        yawn_remaining = max(0.0, YAWN_ALARM_COOLDOWN_SECONDS - (now - self.last_yawn_alarm_at))
        if self.audio_error:
            return self.audio_error
        return (
            f"Drowsy timer {drowsy_seconds:.1f}/{DROWSY_ALARM_DELAY_SECONDS:.1f}s | "
            f"Yawn cooldown {yawn_remaining:.1f}s"
        )


alarm_controller = AlarmController()


class DailyExcelLogger:
    HEADERS = ["Timestamp", "Status", "Confidence", "Session ID", "Device Serial", "Recording File"]

    def __init__(self, path=EXCEL_LOG_PATH):
        self.path = Path(path)
        self.lock = threading.Lock()
        self.warning_shown = False

    def log_status(self, status, confidence=0.0, recording_path=""):
        timestamp = datetime.now()
        with self.lock:
            try:
                from openpyxl import Workbook, load_workbook
            except ImportError:
                if not self.warning_shown:
                    print("Excel logging skipped: install openpyxl with: pip install openpyxl")
                    self.warning_shown = True
                return

            self.path.parent.mkdir(parents=True, exist_ok=True)
            if self.path.exists():
                workbook = load_workbook(self.path)
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)

            sheet_name = timestamp.strftime("%Y-%m-%d")
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
                sheet.append(self.HEADERS)
                self.format_sheet(sheet)

            sheet.append(
                [
                    timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                    status,
                    round(float(confidence or 0.0), 4),
                    DEVICE_SESSION_ID,
                    DEVICE_SERIAL,
                    str(recording_path or ""),
                ]
            )
            workbook.save(self.path)
            self.trim_if_needed(load_workbook)

    def format_sheet(self, sheet):
        widths = [22, 16, 14, 44, 22, 70]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width

    def trim_if_needed(self, load_workbook):
        if not self.path.exists() or self.path.stat().st_size <= EXCEL_MAX_BYTES:
            return

        target_size = max(1024 * 1024, EXCEL_MAX_BYTES - EXCEL_TRIM_BYTES)
        attempts = 0
        while self.path.exists() and self.path.stat().st_size > target_size and attempts < 30:
            attempts += 1
            workbook = load_workbook(self.path)
            trimmed = False

            for sheet_name in sorted(workbook.sheetnames):
                sheet = workbook[sheet_name]
                if sheet.max_row > 1:
                    rows_to_delete = min(max(1, sheet.max_row // 3), sheet.max_row - 1)
                    sheet.delete_rows(2, rows_to_delete)
                    trimmed = True
                    break
                if len(workbook.sheetnames) > 1:
                    del workbook[sheet_name]
                    trimmed = True
                    break

            if not trimmed:
                break
            workbook.save(self.path)

        print(f"Excel log trimmed: {self.path.name} is {self.path.stat().st_size / (1024 * 1024):.1f} MB")


class FatigueEventExcelLogger:
    def __init__(self, excel_logger):
        self.excel_logger = excel_logger
        self.was_yawn = False

    def log_drowsy_confirmation(self, recording_event):
        if not recording_event:
            return
        self.excel_logger.log_status(
            "Drowsy (3 consecutive seconds)",
            recording_event.get("confidence", 0.0),
            recording_path=recording_event.get("recording_path", ""),
        )

    def update_yawn(self, current_status, detections):
        detected_yawn = current_status == "Yawn" or any(
            detection.get("status") == "Yawn" for detection in detections
        )
        if detected_yawn and not self.was_yawn:
            self.excel_logger.log_status("Yawn", detection_confidence("Yawn", detections))
        self.was_yawn = detected_yawn


class DrowsyRecordingManager:
    VIDEO_SUFFIXES = {".mp4", ".avi", ".mov", ".mkv"}

    def __init__(self):
        self.directory = DROWSY_RECORDING_DIR
        self.buffer = deque()
        self.active = False
        self.active_frames = []
        self.active_path = None
        self.active_finish_at = 0.0
        self.active_confidence = 0.0
        self.was_drowsy = False
        self.drowsy_started_at = None
        self.recorded_current_drowsy_event = False
        self.write_threads = []
        self.cleanup_lock = threading.Lock()
        self.directory.mkdir(parents=True, exist_ok=True)

    def update(self, frame, current_status, detections):
        now = time.monotonic()
        self.remember_frame(now, frame)
        detected_drowsy = current_status == "Drowsy" or any(
            detection.get("status") == "Drowsy" for detection in detections
        )

        if detected_drowsy:
            if self.drowsy_started_at is None:
                self.drowsy_started_at = now
        else:
            self.drowsy_started_at = None
            self.recorded_current_drowsy_event = False

        recording_event = None
        drowsy_elapsed = 0.0
        if self.drowsy_started_at is not None:
            drowsy_elapsed = now - self.drowsy_started_at

        if (
            detected_drowsy
            and not self.recorded_current_drowsy_event
            and not self.active
            and drowsy_elapsed >= DROWSY_ALARM_DELAY_SECONDS
        ):
            confidence = detection_confidence("Drowsy", detections)
            recording_event = self.start_recording(now, confidence)
            self.recorded_current_drowsy_event = True

        if self.active and recording_event is None:
            self.active_frames.append((now, frame.copy()))

        if self.active and now >= self.active_finish_at:
            self.finish_recording(wait=False)

        self.was_drowsy = detected_drowsy
        return recording_event

    def remember_frame(self, now, frame):
        self.buffer.append((now, frame.copy()))
        while self.buffer and now - self.buffer[0][0] > DROWSY_RECORDING_PRE_SECONDS:
            self.buffer.popleft()

    def start_recording(self, now, confidence):
        timestamp = datetime.now()
        file_name = f"drowsy_{timestamp.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}.mp4"
        self.active = True
        self.active_path = self.directory / file_name
        self.active_confidence = confidence
        self.active_finish_at = now + max(0.0, DROWSY_RECORDING_SECONDS - DROWSY_RECORDING_PRE_SECONDS)
        self.active_frames = [(frame_time, frame.copy()) for frame_time, frame in self.buffer]
        print(f"Drowsy recording started: {self.active_path.name}")
        return {
            "status": "Drowsy",
            "confidence": confidence,
            "recording_path": str(self.active_path),
            "timestamp": timestamp,
        }

    def finish_recording(self, wait=False):
        if not self.active:
            return

        frames = list(self.active_frames)
        path = self.active_path
        confidence = self.active_confidence
        self.active = False
        self.active_frames = []
        self.active_path = None
        self.active_finish_at = 0.0
        self.active_confidence = 0.0

        if not frames or path is None:
            return

        if wait:
            self.write_video(frames, path, confidence)
        else:
            thread = threading.Thread(target=self.write_video, args=(frames, path, confidence), daemon=True)
            thread.start()
            self.write_threads.append(thread)

    def write_video(self, frames, path, confidence):
        try:
            first_time = frames[0][0]
            last_time = frames[-1][0]
            duration = max(1.0, last_time - first_time)
            fps = max(1.0, min(30.0, len(frames) / duration))
            first_frame = frames[0][1]
            height, width = first_frame.shape[:2]
            codec = (DROWSY_RECORDING_CODEC[:4] or "mp4v").ljust(4)[:4]
            writer = cv2.VideoWriter(str(path), cv2.VideoWriter_fourcc(*codec), fps, (width, height))
            if not writer.isOpened():
                print(f"Could not create drowsy recording: {path}")
                return

            for _, frame in frames:
                if frame.shape[:2] != (height, width):
                    frame = cv2.resize(frame, (width, height))
                writer.write(frame)
            writer.release()
            print(f"Drowsy recording saved: {path} ({len(frames)} frames, {fps:.1f} FPS, confidence {confidence:.2f})")
            self.cleanup_recordings(keep_path=path)
        except Exception as error:
            print(f"Drowsy recording error: {error}")

    def cleanup_recordings(self, keep_path=None):
        with self.cleanup_lock:
            files = [
                path
                for path in self.directory.glob("*")
                if path.is_file() and path.suffix.lower() in self.VIDEO_SUFFIXES
            ]
            total_size = sum(path.stat().st_size for path in files)
            if total_size <= DROWSY_RECORDING_MAX_BYTES:
                return

            keep_resolved = keep_path.resolve() if keep_path else None
            for old_file in sorted(files, key=lambda item: item.stat().st_mtime):
                if keep_resolved and old_file.resolve() == keep_resolved:
                    continue
                try:
                    file_size = old_file.stat().st_size
                    old_file.unlink()
                    total_size -= file_size
                    print(f"Deleted old drowsy recording: {old_file.name}")
                except OSError as error:
                    print(f"Could not delete old recording {old_file.name}: {error}")
                if total_size <= DROWSY_RECORDING_MAX_BYTES:
                    break

    def shutdown(self):
        self.finish_recording(wait=True)
        for thread in self.write_threads:
            thread.join(timeout=8)


excel_logger = DailyExcelLogger()
fatigue_event_excel_logger = FatigueEventExcelLogger(excel_logger)
drowsy_recorder = DrowsyRecordingManager()


@lru_cache(maxsize=1)
def load_model(model_path=str(MODEL_PATH)):
    from ultralytics import YOLO

    return YOLO(model_path)


def class_name(model, class_id):
    names = getattr(model, "names", {})
    if isinstance(names, dict):
        return names.get(class_id, str(class_id))
    try:
        return names[class_id]
    except Exception:
        return str(class_id)


def normalize_status(label):
    text = str(label).strip().lower()
    if "yawn" in text:
        return "Yawn"
    if "drows" in text or "sleep" in text or "closed" in text:
        return "Drowsy"
    if "awake" in text or "open" in text or "alert" in text:
        return "Awake"
    return "No Detection"


def choose_status(statuses):
    if "Yawn" in statuses:
        return "Yawn"
    if "Drowsy" in statuses:
        return "Drowsy"
    if "Awake" in statuses:
        return "Awake"
    return "No Detection"


def utc_timestamp():
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def detection_confidence(status, detections):
    matching = [
        float(detection.get("confidence", 0))
        for detection in detections
        if detection.get("status") == status
    ]
    return max(matching) if matching else 0.0


def send_detection_event(event_type, confidence, status=None, frame_url=None, session_id=None):
    if event_type not in {"Drowsy", "Yawn"}:
        return False
    if not DEVICE_SERIAL or not DEVICE_TOKEN:
        print("Supabase event skipped: set DFMS_DEVICE_SERIAL and DFMS_DEVICE_TOKEN.")
        return False
    if not SUPABASE_EVENT_URL.startswith("https://"):
        print("Supabase event skipped: SUPABASE_EVENT_URL must use HTTPS.")
        return False

    payload = {
        "device_serial": DEVICE_SERIAL,
        "device_token": DEVICE_TOKEN,
        "timestamp": utc_timestamp(),
        "event_type": event_type,
        "confidence": round(float(confidence), 4),
        "status": status or event_type,
        "frame_url": frame_url,
        "session_id": session_id or DEVICE_SESSION_ID,
    }

    request = urllib.request.Request(
        SUPABASE_EVENT_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=EVENT_SEND_TIMEOUT_SECONDS) as response:
            if 200 <= response.status < 300:
                return True
            print(f"Supabase event failed with HTTP {response.status}.")
    except urllib.error.HTTPError as error:
        details = error.read().decode("utf-8", errors="replace")
        print(f"Supabase event rejected: HTTP {error.code} {details}")
    except urllib.error.URLError as error:
        print(f"Supabase event network error: {error.reason}")
    except Exception as error:
        print(f"Supabase event error: {error}")
    return False


def maybe_send_detection_event(status, detections, frame_url=None, confidence_override=None):
    if status not in {"Drowsy", "Yawn"}:
        return

    now = time.monotonic()
    last_sent = _last_event_sent_at.get(status, 0)
    if EVENT_COOLDOWN_SECONDS > 0 and now - last_sent < EVENT_COOLDOWN_SECONDS:
        return

    _last_event_sent_at[status] = now
    confidence = (
        float(confidence_override)
        if confidence_override is not None
        else detection_confidence(status, detections)
    )
    thread = threading.Thread(
        target=send_detection_event,
        args=(status, confidence, status, frame_url, DEVICE_SESSION_ID),
        daemon=True,
    )
    thread.start()


def maybe_send_detection_events(detections, frame_url=None, drowsy_confirmed_event=None):
    if drowsy_confirmed_event:
        maybe_send_detection_event(
            "Drowsy",
            detections,
            frame_url=frame_url,
            confidence_override=drowsy_confirmed_event.get("confidence", 0.0),
        )

    detected_statuses = {detection.get("status") for detection in detections}
    if "Yawn" in detected_statuses:
        maybe_send_detection_event("Yawn", detections, frame_url=frame_url)


def draw_status_banner(frame, status):
    color = COLORS.get(status, COLORS["No Detection"])
    height, width = frame.shape[:2]
    overlay = frame.copy()
    cv2.rectangle(overlay, (0, 0), (width, 66), (12, 38, 85), -1)
    cv2.addWeighted(overlay, 0.68, frame, 0.32, 0, frame)
    cv2.circle(frame, (28, 34), 10, color, -1, cv2.LINE_AA)
    cv2.putText(
        frame,
        f"Status: {status}",
        (48, 42),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.85,
        (255, 255, 255),
        2,
        cv2.LINE_AA,
    )


def draw_alarm_status(frame, alarm_text):
    height, width = frame.shape[:2]
    text = f"Alarm: {alarm_text}"
    y = min(96, max(78, height - 16))
    cv2.rectangle(frame, (0, 66), (width, 112), (12, 38, 85), -1)
    cv2.putText(
        frame,
        text,
        (18, y),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.58,
        (219, 234, 254),
        2,
        cv2.LINE_AA,
    )
    cv2.putText(
        frame,
        time.strftime("%Y-%m-%d %H:%M:%S"),
        (max(width - 285, 48), 42),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.65,
        (219, 234, 254),
        2,
        cv2.LINE_AA,
    )


def detect_fatigue(frame, confidence=0.5):
    model = load_model()
    inference_options = {"conf": confidence, "verbose": False}
    if MODEL_IMGSZ > 0:
        inference_options["imgsz"] = MODEL_IMGSZ
    results = model(frame, **inference_options)
    result = results[0]
    annotated = frame.copy()
    statuses = []
    detections = []

    for box in result.boxes:
        class_id = int(box.cls[0])
        raw_label = class_name(model, class_id)
        status = normalize_status(raw_label)
        x1, y1, x2, y2 = box.xyxy[0].cpu().numpy().astype(int)
        conf = float(box.conf[0])
        if status == "Yawn" and conf < YAWN_MIN_CONFIDENCE:
            status = "No Detection"
        statuses.append(status)

        color = COLORS.get(status, (255, 255, 255))
        label = f"{status} {conf:.2f}"

        cv2.rectangle(annotated, (x1, y1), (x2, y2), color, 3)
        cv2.putText(
            annotated,
            label,
            (x1, max(y1 - 12, 24)),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            color,
            2,
            cv2.LINE_AA,
        )
        detections.append(
            {
                "status": status,
                "raw_label": str(raw_label),
                "confidence": conf,
                "box": (x1, y1, x2, y2),
            }
        )

    current_status = choose_status(statuses)
    draw_status_banner(annotated, current_status)
    return annotated, current_status, detections


def run_camera(camera_index=0, confidence=0.5):
    print(f"Device serial: {DEVICE_SERIAL or 'not configured'}")
    print(f"Device token: {'configured' if DEVICE_TOKEN else 'not configured'}")
    print(f"Supabase endpoint: {SUPABASE_EVENT_URL}")
    print(f"Device config file: {DEVICE_CONFIG_PATH}")
    cap = cv2.VideoCapture(camera_index)
    if not cap.isOpened():
        raise RuntimeError(f"Could not open camera index {camera_index}.")
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, CAMERA_WIDTH)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, CAMERA_HEIGHT)
    alarm_controller.ensure_audio()

    try:
        while True:
            ok, frame = cap.read()
            if not ok:
                break
            annotated, current_status, detections = detect_fatigue(frame, confidence=confidence)
            alarm_controller.update(current_status, detections)
            draw_alarm_status(annotated, alarm_controller.last_status)
            recording_event = drowsy_recorder.update(annotated, current_status, detections)
            fatigue_event_excel_logger.log_drowsy_confirmation(recording_event)
            fatigue_event_excel_logger.update_yawn(current_status, detections)
            maybe_send_detection_events(detections, drowsy_confirmed_event=recording_event)
            cv2.imshow("YOLO Fatigue Detection", annotated)
            key = cv2.waitKey(1) & 0xFF
            if key == ord("1"):
                alarm_controller.test_drowsy_alarm()
            elif key == ord("2"):
                alarm_controller.test_yawn_alarm()
            elif key == ord("0"):
                alarm_controller.stop_all()
                alarm_controller.ensure_audio()
            elif key in (ord("q"), 27):
                break
    finally:
        drowsy_recorder.shutdown()
        alarm_controller.stop_all()
        cap.release()
        cv2.destroyAllWindows()


def test_alarms():
    print(f"Testing {DROWSY_ALARM_PATH.name} for 3 seconds...")
    alarm_controller.test_drowsy_alarm()
    time.sleep(3)
    alarm_controller.stop_all()
    print(f"Testing {YAWN_ALARM_PATH.name} for 3 seconds...")
    alarm_controller.test_yawn_alarm()
    time.sleep(3)
    alarm_controller.stop_all()
    if alarm_controller.audio_error:
        print(f"Alarm test failed: {alarm_controller.audio_error}")
    else:
        print("Alarm test finished. If you heard nothing, check Windows output device/volume mixer.")


def setup_device_config():
    print("One-time device setup")
    print("This saves the device serial and token locally on this physical device.")
    serial = input("Device serial number: ").strip()
    token = input("Device token: ").strip()
    default_url = "https://wxxdtuzicvjkjqymjsfj.supabase.co/functions/v1/device-events"
    event_url = input(f"Supabase event URL [{default_url}]: ").strip() or default_url

    if not serial or not token:
        raise RuntimeError("Device serial and device token are required.")
    if not event_url.startswith("https://"):
        raise RuntimeError("Supabase event URL must start with https://")

    config = {
        "device_serial": serial,
        "device_token": token,
        "supabase_event_url": event_url,
    }
    DEVICE_CONFIG_PATH.write_text(json.dumps(config, indent=2), encoding="utf-8")
    print(f"Saved device config to: {DEVICE_CONFIG_PATH}")
    print("Next time, run only: python main.py")


if __name__ == "__main__":
    if "--setup-device" in sys.argv:
        setup_device_config()
    elif "--test-alarms" in sys.argv:
        test_alarms()
    else:
        run_camera()
